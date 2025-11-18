import requests
from fastapi import FastAPI, Body
import uvicorn
from openpyxl import Workbook
import io
from fastapi.responses import StreamingResponse
from urllib.parse import quote


app = FastAPI()

class UsersInformation():
    def __init__(self, users_list = [], users_id = []):
        self.users_list = users_list or []
        self.users_id = users_id or []

    def user_id(self):
        if self.users_list != []:
            for user in self.users_list:                   # Потом убрать
                self.users_id.append(user["id"])
            return self.users_id
        else:
            return None

    """def Login_Func(self):
        url = 'http://intranet.emk.org.ru:8000/api/auth_router/auth'
        payload = {'login': 'bragin.i.d@emk.ru', 'password': 'XC74JSnMW5'}
        # payload = {'login': 'kucherenko.m.d@emk.ru', 'password': 'RyfcfRF7vG'}
        dataD = json.dumps(payload)
        response = requests.post(url, data=dataD, timeout=120)
        # print(response.headers)
        dataL = json.loads(response.text)
        if 'status' in dataL and dataL['status'] == 'error':
            return False
        else:
            return True"""

    def get_requests(self):
        users_info = []
        users_id = self.user_id()
        headers = {'Authorization': 'a9900b46-cb29-4b28-9a11-e5bbca8c2a45'}
        for id in users_id:
            response = requests.get(f'http://intranet.emk.org.ru:8000/api/users/find_by/{id}', headers = headers)
            users_info.append(response.json())
            # print(id)
            # print(response.status_code)
            # print(response.text)
            # print(response.headers)
            # return response.text
        return users_info



    """def read_users(self):
        try:
            with open("./users.json", "r", encoding="utf-8") as pattern_data_file:
                isload = json.load(pattern_data_file)
            return isload
        except:
            return []"""

    def make_excel_list(self):
        users_info = self.get_requests()

        # Создание нового файла
        wb = Workbook()
        ws = wb.active
        ws.title = "Список участников"

        # Запись данных
        ws['A1'] = 'ФИО'
        ws['B1'] = 'EMAIL'
        ws['C1'] = 'Телефон'
        ws['D1'] = 'Внутр. номер'
        ws['E1'] = 'Должность'
        ws['F1'] = 'Дирекция/Завод'
        ws['G1'] = 'Подразделение'
        ws['H1'] = 'Местоположение'

        for i, user in enumerate(users_info, start=2):  # users это список словарей из данных о пользователях. То есть
            # переделать, чтобы посылать запрос о информации о пользователе и уже этот текст обрабатывать в этой функции
            indirect_data = user.get("indirect_data", {})

            if "name" in user and "last_name" in user and "second_name" in user: ws[
                f'A{i}'] = f'{user["name"]} {user["last_name"]} {user["second_name"]}'
            if "email" in user: ws[f'B{i}'] = f'{user["email"]}'
            if "personal_mobile" in user: ws[f'C{i}'] = f'{user["personal_mobile"]}'
            if "uf_phone_inner" in user: ws[f'D{i}'] = f'{user["uf_phone_inner"]}'
            if "work_position" in indirect_data:
                ws[f'E{i}'] = f'{indirect_data["work_position"]}'
            if "uf_department" in indirect_data and "uf_usr_1696592324977" in indirect_data:
                ws[f'F{i}'] = ", ".join(indirect_data["uf_usr_1696592324977"]) + "/" + ", ".join(
                    indirect_data["uf_department"])
            if "uf_usr_1705744824758" in indirect_data:
                ws[f'G{i}'] = " ".join(indirect_data["uf_usr_1705744824758"])
            if "personal_city" in user: ws[f'H{i}'] = f'{user["personal_city"]}'

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Сохранение
        return excel_buffer



@app.post("/EventsInExcel", summary = "Скачать Excel со всеми участниками мероприятия")
def make_users_excel_list(data: list = Body()):
    excel_buffer = UsersInformation(users_list = data).make_excel_list()

    return StreamingResponse(excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers = {"Content-Disposition": "attachment; filename=participants.xlsx"})
    # return UsersInformation().Login_Func()
    # return UsersInformation(users_list = data).get_requests()


if __name__ == '__main__':
    uvicorn.run(app, host="127.0.0.1", port=8000)